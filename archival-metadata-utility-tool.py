import os
import pandas as pd
from datetime import datetime
import argparse
import difflib
import mimetypes 
import urllib.parse 
import sys
import threading
import queue
import shutil
import re
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext

from openpyxl.utils import get_column_letter

print("\n--- Running Archival Metadata Utility Tool ---")

def sanitize_filename(name):
    """
    Sanitizes string to be safe for filenames.
    """
    clean = re.sub(r'[\\/*?:"<>|]', '', str(name)).strip()
    return clean if clean else "Archival"

def get_dcmi_type(mime):
    """
    Maps MIME types strictly to canonical DCMI Type Vocabulary terms.
    https://www.dublincore.org/specifications/dublin-core/dcmi-type-vocabulary/
    """
    mime_lower = str(mime).lower().strip()
    
    if mime_lower == 'inode/directory':
        return 'Collection'
    
    dataset_apps = ['csv', 'excel', 'spreadsheet', 'spreadsheetml']
    if any(d in mime_lower for d in dataset_apps):
        return 'Dataset'
        
    text_apps = ['pdf', 'msword', 'wordprocessingml', 'rtf', 'epub', 'document']
    if any(t in mime_lower for t in text_apps):
        return 'Text'
        
    collection_apps = ['zip', 'tar', 'compressed', 'rar']
    if any(c in mime_lower for c in collection_apps):
        return 'Collection'
        
    software_apps = ['executable', 'software']
    if any(s in mime_lower for s in software_apps):
        return 'Software'

    if mime_lower.startswith('image/'):
        return 'Image'
    elif mime_lower.startswith('audio/'):
        return 'Sound'
    elif mime_lower.startswith('video/'):
        return 'MovingImage'
    elif mime_lower.startswith('text/'):
        return 'Text'
        
    return 'Dataset'

def extract_raw_metadata(source_folder):
    """
    Extracts raw technical metadata natively, recursively traversing nested folders.
    """
    print(f"Extracting metadata from: {source_folder}")
    metadata_list = []
    
    norm_path = os.path.normpath(source_folder)
    top_level_folder = os.path.basename(norm_path)
    if not top_level_folder:
        top_level_folder = norm_path.rstrip('\\/').rstrip(':') or source_folder

    def get_creation_date(stats):
        # Prefer st_birthtime (macOS creation time) if available, fallback to st_ctime
        ctime = getattr(stats, 'st_birthtime', stats.st_ctime)
        return datetime.fromtimestamp(ctime).strftime('%Y-%m-%d')

    for root, dirs, files in os.walk(source_folder):
        # Exclude hidden directories in-place so os.walk does not traverse into them
        dirs[:] = [d for d in dirs if not d.startswith('.')]
        
        # 1. Add current folder entry
        try:
            folder_stats = os.stat(root)
            folder_name = os.path.basename(root) if root != norm_path else top_level_folder
            if not folder_name:
                folder_name = top_level_folder
                
            metadata_list.append({
                "Top Level Folder Name": top_level_folder,
                "File_Name": folder_name,
                "File Path": root,
                "Size (Bytes)": folder_stats.st_size,
                "Creation_Date": get_creation_date(folder_stats),
                "Format_Type": "folder",
                "Author_Creator": ""
            })
        except Exception as e:
            print(f"Warning: Could not stat directory {root}: {e}")

        # 2. Add files in current folder
        for file in files:
            if file.startswith('.'):
                continue
                
            file_path = os.path.join(root, file)
            try:
                stats = os.stat(file_path)
                metadata_list.append({
                    "Top Level Folder Name": top_level_folder,
                    "File_Name": file,
                    "File Path": file_path,
                    "Size (Bytes)": stats.st_size,
                    "Creation_Date": get_creation_date(stats),
                    "Format_Type": os.path.splitext(file)[1].lower().strip(),
                    "Author_Creator": ""
                })
            except Exception as e:
                print(f"Warning: Could not stat file {file_path}: {e}")
            
    return pd.DataFrame(metadata_list)

def generate_opex_xml(completed_dc_df, output_dir, ref_folder_name=""):
    """
    Generates Preservica OPEX XML metadata file(s) from validated Dublin Core metadata.
    """
    import xml.etree.ElementTree as ET
    from xml.dom import minidom

    headers = [str(col).strip() for col in completed_dc_df.iloc[1].tolist()]
    data_rows = completed_dc_df.iloc[2:]

    dc_tag_map = {
        "title": "dc:title",
        "creator": "dc:creator",
        "subject": "dc:subject",
        "description": "dc:description",
        "publisher": "dc:publisher",
        "contributor": "dc:contributor",
        "date": "dc:date",
        "type": "dc:type",
        "format": "dc:format",
        "identifier": "dc:identifier",
        "source": "dc:source",
        "language": "dc:language",
        "relation": "dc:relation",
        "coverage": "dc:coverage",
        "rights": "dc:rights"
    }

    fn_col_idx = None
    for idx, h in enumerate(headers):
        if h.lower() in ["file or folder name", "file name"]:
            fn_col_idx = idx
            break

    opex_files_created = []

    # Master OPEX XML file for the dataset/collection
    master_opex = ET.Element("opex:OPEXMetadata", {
        "xmlns:opex": "http://www.openpreservationexchange.org/opex/v1.0"
    })
    descriptive = ET.SubElement(master_opex, "opex:DescriptiveMetadata")

    for r_idx, row in data_rows.iterrows():
        oai_dc = ET.SubElement(descriptive, "oai_dc:dc", {
            "xmlns:oai_dc": "http://www.openarchives.org/OAI/2.0/oai_dc/",
            "xmlns:dc": "http://purl.org/dc/elements/1.1/",
            "xmlns:xsi": "http://www.w3.org/2001/XMLSchema-instance"
        })

        row_fn = ""
        if fn_col_idx is not None:
            val_fn = row.iloc[fn_col_idx]
            if pd.notna(val_fn):
                row_fn = str(val_fn).strip()

        # Sidecar item OPEX XML
        item_opex = ET.Element("opex:OPEXMetadata", {
            "xmlns:opex": "http://www.openpreservationexchange.org/opex/v1.0"
        })
        item_desc = ET.SubElement(item_opex, "opex:DescriptiveMetadata")
        item_oai_dc = ET.SubElement(item_desc, "oai_dc:dc", {
            "xmlns:oai_dc": "http://www.openarchives.org/OAI/2.0/oai_dc/",
            "xmlns:dc": "http://purl.org/dc/elements/1.1/",
            "xmlns:xsi": "http://www.w3.org/2001/XMLSchema-instance"
        })

        has_item_data = False

        for c_idx, col_name in enumerate(headers):
            val = row.iloc[c_idx]
            if pd.isna(val):
                continue
            val_str = str(val).strip()
            if not val_str or val_str.lower() in ("nan", "none"):
                continue

            col_lower = col_name.lower()
            if col_lower in dc_tag_map:
                tag_name = dc_tag_map[col_lower]
                
                elem = ET.SubElement(oai_dc, tag_name)
                elem.text = val_str
                
                item_elem = ET.SubElement(item_oai_dc, tag_name)
                item_elem.text = val_str
                has_item_data = True
            elif col_lower in ["entity ref", "entity reference id"]:
                elem = ET.SubElement(oai_dc, "opex:EntityRef")
                elem.text = val_str
                item_elem = ET.SubElement(item_oai_dc, "opex:EntityRef")
                item_elem.text = val_str
                has_item_data = True

        if row_fn and has_item_data:
            sidecar_filename = f"{row_fn}.opex"
            sidecar_path = os.path.join(output_dir, sidecar_filename)
            try:
                raw_bytes = ET.tostring(item_opex, encoding="utf-8")
                xml_str = minidom.parseString(raw_bytes).toprettyxml(indent="  ")
                xml_str = "\n".join([line for line in xml_str.splitlines() if line.strip()])
                with open(sidecar_path, "w", encoding="utf-8") as f:
                    f.write(xml_str + "\n")
                opex_files_created.append(sidecar_path)
            except Exception as e:
                print(f"Warning: Could not write sidecar OPEX file {sidecar_path}: {e}")

    master_filename = f"{ref_folder_name}.opex" if ref_folder_name else "DC.opex"
    master_path = os.path.join(output_dir, master_filename)
    try:
        raw_bytes = ET.tostring(master_opex, encoding="utf-8")
        xml_str = minidom.parseString(raw_bytes).toprettyxml(indent="  ")
        xml_str = "\n".join([line for line in xml_str.splitlines() if line.strip()])
        with open(master_path, "w", encoding="utf-8") as f:
            f.write(xml_str + "\n")
        opex_files_created.append(master_path)
        print(f"Success! Exported OPEX metadata XML to {master_path}")
    except Exception as e:
        print(f"Error writing master OPEX XML file {master_path}: {e}")

    return opex_files_created

def resolve_template_path(template_path=None):
    """
    Resolves template_path location reliably across different CWDs, script locations, and alternate template names.
    """
    base_dir = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))

    if template_path and os.path.exists(template_path):
        return os.path.abspath(template_path)

    candidates = []
    if template_path:
        candidates.append(template_path)
        candidates.append(os.path.join(base_dir, template_path))

    for name in ["Metadata_Template.xlsx", "Working_Metadata.xlsx"]:
        candidates.append(os.path.join(base_dir, name))
        candidates.append(name)

    for cand in candidates:
        if cand and os.path.exists(cand):
            return os.path.abspath(cand)

    return template_path or os.path.join(base_dir, "Metadata_Template.xlsx")

def process_metadata(source_folder=None, template_path="Metadata_Template.xlsx", preservica_csv=None, existing_metadata=None, output_dir=None, on_pause_callback=None, resume=False, on_resume_callback=None, workflow=None, opex_export=False):
    """
    Processes raw metadata, copies template_path to <output_dir>/<Folder_Name>_Metadata.xlsx,
    merges data into the copy, validates against controlled vocabulary, and outputs DC.csv in output_dir.
    Supports resuming existing metadata workbooks without overwriting manual DC edits.
    """
    if workflow is None:
        if preservica_csv:
            workflow = 3 if existing_metadata else 2
        elif resume or (existing_metadata and not source_folder):
            workflow = 4
        else:
            workflow = 1
    if not output_dir or os.path.abspath(output_dir) == os.path.abspath("."):
        print("Error: Export directory cannot be the root directory. Please specify a different export output directory.")
        return

    template_path = resolve_template_path(template_path)
    if not os.path.exists(template_path):
        print(f"Error: Cannot find the mapping workbook template at {template_path}")
        return

    # Check if we are running Option 3/4 (existing metadata file) or extracting from a directory
    if existing_metadata:
        print(f"Loading existing technical metadata from: {existing_metadata}")
        if str(existing_metadata).endswith('.csv'):
            raw_df = pd.read_csv(existing_metadata)
        else:
            try:
                raw_df = pd.read_excel(existing_metadata, sheet_name="Raw_Data")
            except ValueError:
                try:
                    raw_df = pd.read_excel(existing_metadata, sheet_name="Raw Data")
                except ValueError:
                    raw_df = pd.read_excel(existing_metadata)
    elif source_folder:
        raw_df = extract_raw_metadata(source_folder)
    else:
        raw_df = pd.DataFrame()
    
    if raw_df.empty and not resume:
        print("Warning: No files/records found.")
        return

    # Determine referenced folder or device name for naming output Excel workbook
    ref_folder_name = ""
    if source_folder:
        norm_p = os.path.normpath(source_folder)
        ref_folder_name = os.path.basename(norm_p) or norm_p.rstrip('\\/').rstrip(':')

    if not ref_folder_name and "Top Level Folder Name" in raw_df.columns and not raw_df["Top Level Folder Name"].dropna().empty:
        ref_folder_name = str(raw_df["Top Level Folder Name"].dropna().iloc[0]).strip()

    if not ref_folder_name and existing_metadata:
        ref_folder_name = os.path.splitext(os.path.basename(existing_metadata))[0]

    ref_folder_name = sanitize_filename(ref_folder_name)
    if ref_folder_name.lower().endswith("_metadata"):
        ref_folder_name = ref_folder_name[:-9].strip()

    target_excel_filename = f"{ref_folder_name}_Metadata.xlsx" if ref_folder_name else "Metadata.xlsx"

    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
        excel_path = os.path.join(output_dir, target_excel_filename)
    else:
        excel_path = target_excel_filename

    is_resuming = False
    if os.path.exists(excel_path):
        try:
            test_dc = pd.read_excel(excel_path, sheet_name="DC", header=None)
            if not test_dc.empty:
                if resume:
                    is_resuming = True
                elif on_resume_callback:
                    is_resuming = on_resume_callback(excel_path)
                elif sys.stdin and sys.stdin.isatty():
                    ans = input(f"\nFound existing metadata workbook at '{excel_path}'. Resume using existing file? (y/n): ").strip().lower()
                    if ans in ('y', 'yes'):
                        is_resuming = True
        except Exception:
            is_resuming = False

    if is_resuming:
        print(f"\n[RESUME DETECTED] Using existing metadata workbook: {excel_path}")
        print("Preserving existing DC edits...")
    else:
        if os.path.abspath(template_path) != os.path.abspath(excel_path):
            print(f"Creating metadata workbook copy: {excel_path} from template: {template_path}")
            shutil.copy2(template_path, excel_path)
        else:
            print(f"Using template directly at: {excel_path}")

        print(f"Writing extracted/loaded raw data to {excel_path}...")
        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            raw_df.to_excel(writer, sheet_name="Raw Data", index=False)

    if not is_resuming:
        print("Mapping to Dublin Core standards...")
        try:
            format_types_df = pd.read_excel(excel_path, sheet_name="Format_Types", header=None)
            existing_dc_df = pd.read_excel(excel_path, sheet_name="DC", header=None)
            dc_headers = existing_dc_df.iloc[0:2].copy() 
        except Exception as e:
            print(f"Error reading tabs from Excel: {e}")
            return
            
        template_columns = [str(col).strip() for col in dc_headers.iloc[1].tolist()]
        
        if workflow == 1:
            cols_to_remove = [idx for idx, c in enumerate(template_columns) if str(c).strip().lower() in ["entity ref", "entity reference id"]]
            for idx in sorted(cols_to_remove, reverse=True):
                template_columns.pop(idx)
                dc_headers = dc_headers.drop(dc_headers.columns[idx], axis=1)
            dc_headers.columns = range(dc_headers.shape[1])

        template_cols_lower = {str(c).lower(): str(c) for c in template_columns}
        
        col_file = template_cols_lower.get("file or folder name", "File or Folder Name")
        col_creator = template_cols_lower.get("creator", "Creator")
        col_date = template_cols_lower.get("date", "Date")
        col_format = template_cols_lower.get("format", "Format")
        col_type = template_cols_lower.get("type", "Type")

        format_mapping = {}
        dcmi_mapping = {}
        if format_types_df.shape[1] >= 5:
            for ext_val, mime_val, dcmi_val in zip(format_types_df.iloc[:, 2], format_types_df.iloc[:, 3], format_types_df.iloc[:, 4]):
                if pd.notna(ext_val) and str(ext_val).strip() != "":
                    ext_str = str(ext_val).strip()
                    ext_list = [e.strip().lower().replace('.', '') for e in ext_str.split(',') if e.strip()]
                    for e in ext_list:
                        if e and e != 'migration':
                            clean_k = "." + e
                            if pd.notna(mime_val) and str(mime_val).strip() != "":
                                format_mapping[clean_k] = str(mime_val).strip()
                            if pd.notna(dcmi_val) and str(dcmi_val).strip() != "":
                                dcmi_str = str(dcmi_val).strip()
                                if dcmi_str == "StillImage":
                                    dcmi_str = "Image"
                                dcmi_mapping[clean_k] = dcmi_str

        new_rows = []
        for _, row in raw_df.iterrows():
            file_name = str(row.get("File_Name", "")).strip()
            raw_format_type = str(row.get("Format_Type", "")).strip()
            
            if raw_format_type.lower() in ["folder", "[root directory]", "directory"]:
                ext = "folder"
            else:
                _, ext = os.path.splitext(file_name)
                ext = ext.lower().strip()
                if not ext or ext == ".":
                    ft_lower = raw_format_type.lower()
                    if "tiff" in ft_lower or "tif" in ft_lower: ext = ".tif"
                    elif "jpeg" in ft_lower or "jpg" in ft_lower: ext = ".jpg"
                    elif "png" in ft_lower: ext = ".png"
                    elif "gif" in ft_lower: ext = ".gif"
                    elif "eps" in ft_lower: ext = ".eps"
                    elif "pdf" in ft_lower: ext = ".pdf"
                    elif "stuffit" in ft_lower: ext = ".sit"
                    elif "word" in ft_lower: ext = ".doc"
                    elif "excel" in ft_lower: ext = ".xls"
                    elif "executable" in ft_lower: ext = ".exe"
                    else: ext = "." + ft_lower.split()[0] if ft_lower else ""

            if ext == "folder":
                mime = "inode/directory"
                dcmi_type = "Collection"
            else:
                mime = format_mapping.get(ext)
                dcmi_type = dcmi_mapping.get(ext)
                
                if not mime:
                    mime_guess, _ = mimetypes.guess_type(file_name)
                    mime = mime_guess if mime_guess else "application/octet-stream"
                    
                if not dcmi_type or dcmi_type in ["Dataset", "StillImage"]:
                    dcmi_type = get_dcmi_type(mime)
                
            raw_date = row.get("Creation_Date", "")
            iso_date = ""
            if pd.notna(raw_date) and str(raw_date).strip():
                parsed_d = pd.to_datetime(str(raw_date).strip(), errors='coerce')
                if pd.notna(parsed_d):
                    iso_date = parsed_d.strftime('%Y-%m-%d')
                else:
                    iso_date = str(raw_date).strip()

            new_rows.append({
                col_file: row.get("File_Name", ""),
                col_creator: row.get("Author_Creator", ""),
                col_date: iso_date,
                col_format: mime,
                col_type: dcmi_type
            })
            
        dc_data_df = pd.DataFrame(new_rows)
                
        if preservica_csv and os.path.exists(preservica_csv):
            print(f"Merging Preservica data from {preservica_csv}...")
            if str(preservica_csv).lower().endswith(('.xlsx', '.xls')):
                preservica_df = pd.read_excel(preservica_csv)
            else:
                preservica_df = pd.read_csv(preservica_csv)
            
            # Find file name column in Preservica export (prioritizing File Name variations)
            file_col_in_preservica = None
            for col in ["File Name", "File_Name", "File or Folder Name", "Title", "Name"]:
                if col in preservica_df.columns:
                    file_col_in_preservica = col
                    break
                    
            # Find Entity Ref column in Preservica export
            entity_ref_col_in_preservica = None
            for col in ["Entity Ref", "Entity Reference ID", "Entity_Ref", "EntityRef"]:
                if col in preservica_df.columns:
                    entity_ref_col_in_preservica = col
                    break
            
            if file_col_in_preservica and entity_ref_col_in_preservica:
                print(f"Matching using Preservica column '{file_col_in_preservica}' and Entity Ref column '{entity_ref_col_in_preservica}'...")
                
                def get_lookup_keys(s):
                    s_str = str(s).strip()
                    unq = urllib.parse.unquote(s_str)
                    base = os.path.splitext(s_str.rstrip('.'))[0]
                    unq_base = os.path.splitext(unq.rstrip('.'))[0]
                    return {s_str.lower(), s_str.rstrip('.').lower(), base.lower(), unq.lower(), unq.rstrip('.').lower(), unq_base.lower()}

                preservica_map = {}
                for _, p_row in preservica_df.iterrows():
                    ref_val = p_row[entity_ref_col_in_preservica]
                    title_val = p_row[file_col_in_preservica]
                    if pd.notna(ref_val) and str(ref_val).strip():
                        for k in get_lookup_keys(title_val):
                            if k not in preservica_map:
                                preservica_map[k] = str(ref_val).strip()

                entity_refs = []
                matched_count = 0
                for r_idx, dc_row in dc_data_df.iterrows():
                    excel_row = r_idx + 3  # Data rows in DC sheet start at row 3
                    fn = str(dc_row.get(col_file, "")).strip()
                    matched_ref = ""
                    for k in get_lookup_keys(fn):
                        if k in preservica_map:
                            matched_ref = preservica_map[k]
                            matched_count += 1
                            break
                    entity_refs.append(matched_ref)
                    if not matched_ref:
                        c_idx = template_columns.index(col_file) if col_file in template_columns else 0
                        col_letter = get_column_letter(c_idx + 1)
                        print(f"[WARNING in Cell {col_letter}{excel_row}] Could not match Preservica Entity Ref ID for asset '{fn}'.")

                print(f"Successfully matched {matched_count} out of {len(dc_data_df)} records with Preservica Entity Ref IDs.")
                dc_data_df["Entity Ref"] = entity_refs
                
                if "Entity Ref" not in template_columns:
                    template_columns.append("Entity Ref")
                    new_col_idx = dc_headers.shape[1]
                    dc_headers.insert(new_col_idx, new_col_idx, ["", "Entity Ref"])
                    dc_headers.columns = range(dc_headers.shape[1])
            else:
                missing_cols = []
                if not file_col_in_preservica:
                    missing_cols.append("File Name / Title")
                if not entity_ref_col_in_preservica:
                    missing_cols.append("Entity Ref / Entity Reference ID")
                print(f"Error: {preservica_csv} is missing required column(s): {', '.join(missing_cols)}.")
                return

        for col in template_columns:
            if col not in dc_data_df.columns:
                dc_data_df[col] = "" 
                
        dc_data_df = dc_data_df[template_columns]
        dc_data_df.columns = dc_headers.columns 
        
        final_dc_export = pd.concat([dc_headers, dc_data_df], ignore_index=True)

        with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            final_dc_export.to_excel(writer, sheet_name="DC", index=False, header=False)
        
    print("\n" + "="*70)
    pause_msg = (f"ACTION REQUIRED: Descriptive metadata must be manually inputted.\n"
                 f"Please open {excel_path}, complete the 'DC' tab, save, and close the file.")
    if on_pause_callback:
        print(pause_msg)
        print("Waiting for user confirmation in GUI to validate and export...")
        on_pause_callback(pause_msg)
    else:
        print(pause_msg + "\nPress ENTER when you are ready to validate and export...")
        input()
    print("="*70 + "\n")

    print("Loading manually entered metadata for validation...")
    try:
        completed_dc_df = pd.read_excel(excel_path, sheet_name="DC", header=None)
        controlled_vocab_df = pd.read_excel(excel_path, sheet_name="Controlled_Vocabulary")
    except Exception as e:
        print(f"Error reading tabs during validation: {e}")
        return

    vocab_list = controlled_vocab_df.iloc[:, 0].dropna().astype(str).str.replace('\xa0', ' ').str.strip().tolist()
    vocab_map = {v.lower(): v for v in vocab_list}
    
    # Sort vocabulary by word count descending to prioritize multi-word phrases
    sorted_vocab_keys = sorted(vocab_map.keys(), key=lambda k: len(k.split()), reverse=True)

    def validate_and_correct_text(text, cell_ref=""):
        if pd.isna(text) or not str(text).strip(): 
            return text, True
            
        text_str = str(text).replace('\xa0', ' ').strip()
        matched_any = False
        
        for vocab_lower in sorted_vocab_keys:
            vocab_correct = vocab_map[vocab_lower]
            vocab_spaceless = vocab_lower.replace(' ', '')
            
            words = text_str.split()
            if not words:
                continue
                
            found_match = False
            for n in range(1, len(words) + 1):
                if found_match:
                    break
                for i in range(len(words) - n + 1):
                    raw_chunk = ' '.join(words[i:i+n])
                    clean_chunk = raw_chunk.strip('.,;:!?\"\'()')
                    if not clean_chunk:
                        continue
                    
                    chunk_spaceless = clean_chunk.lower().replace(' ', '')
                    
                    cell_info = f" in Cell {cell_ref}" if cell_ref else ""
                    
                    if chunk_spaceless == vocab_spaceless:
                        if clean_chunk != vocab_correct:
                            text_str = text_str.replace(clean_chunk, vocab_correct)
                            print(f"[TYPO DETECTED & FIXED{cell_info}] User input: '{clean_chunk}' -> Corrected to: '{vocab_correct}'")
                        found_match = True
                        matched_any = True
                        break
                    
                    sim = difflib.SequenceMatcher(None, clean_chunk.lower(), vocab_lower).ratio()
                    sim_spaceless = difflib.SequenceMatcher(None, chunk_spaceless, vocab_spaceless).ratio()
                    
                    if sim > 0.85 or (len(vocab_spaceless) >= 5 and sim_spaceless > 0.85):
                        if clean_chunk != vocab_correct:
                            text_str = text_str.replace(clean_chunk, vocab_correct)
                            print(f"[TYPO DETECTED & FIXED{cell_info}] User input: '{clean_chunk}' -> Corrected to: '{vocab_correct}'")
                        found_match = True
                        matched_any = True
                        break

        return text_str.strip(), matched_any

    data_rows = completed_dc_df.iloc[2:].astype(object).copy()
    headers = [str(col).strip() for col in completed_dc_df.iloc[1].tolist()]

    for r_idx in data_rows.index:
        excel_row = r_idx + 1  # 1-indexed row number in Excel
        
        # 1. Controlled Vocabulary validation for descriptive text fields
        for field in ["Title", "Description", "Subject", "Contributor", "Creator"]:
            if field in headers:
                c_idx = headers.index(field)
                col_letter = get_column_letter(c_idx + 1)
                cell_ref = f"{col_letter}{excel_row}"
                val = data_rows.loc[r_idx, data_rows.columns[c_idx]]
                if pd.notna(val) and str(val).strip():
                    corrected_val, matched_any = validate_and_correct_text(val, cell_ref=cell_ref)
                    data_rows.loc[r_idx, data_rows.columns[c_idx]] = corrected_val
                    if not matched_any and sorted_vocab_keys and field in ["Title", "Description"]:
                        print(f"[WARNING in Cell {cell_ref}] Field '{field}' value '{str(val).strip()}' does not match controlled vocabulary.")

        # 2. Required field check ('File or Folder Name' / 'File Name')
        fn_c_idx = None
        for fn_col in ["File or Folder Name", "File Name"]:
            if fn_col in headers:
                fn_c_idx = headers.index(fn_col)
                break
        if fn_c_idx is not None:
            col_letter = get_column_letter(fn_c_idx + 1)
            cell_ref = f"{col_letter}{excel_row}"
            val = data_rows.loc[r_idx, data_rows.columns[fn_c_idx]]
            if pd.isna(val) or not str(val).strip():
                print(f"[WARNING in Cell {cell_ref}] Required metadata field '{headers[fn_c_idx]}' is missing or empty.")

        # 3. Date field validation
        if "Date" in headers:
            date_c_idx = headers.index("Date")
            col_letter = get_column_letter(date_c_idx + 1)
            cell_ref = f"{col_letter}{excel_row}"
            val = data_rows.loc[r_idx, data_rows.columns[date_c_idx]]
            if pd.notna(val) and str(val).strip() not in ("", "nan", "None"):
                str_val = str(val).strip()
                parsed_d = pd.to_datetime(str_val, errors='coerce')
                if pd.notna(parsed_d):
                    data_rows.loc[r_idx, data_rows.columns[date_c_idx]] = parsed_d.strftime('%Y-%m-%d')
                else:
                    print(f"[ERROR in Cell {cell_ref}] Invalid date value '{str_val}'. Expected format: YYYY-MM-DD.")
    
    # Filter Entity Ref column for DC export based on workflow option:
    # - Option 2 or 3: export Entity Ref column
    # - Option 4: export Entity Ref column ONLY if there is at least one non-empty value in it
    # - Option 1 / others: do not export Entity Ref column
    entity_ref_col_idx = None
    for idx, col_name in enumerate(headers):
        if str(col_name).strip().lower() in ["entity ref", "entity reference id"]:
            entity_ref_col_idx = idx
            break

    should_export_entity_ref = False
    if entity_ref_col_idx is not None:
        if workflow in (2, 3):
            should_export_entity_ref = True
        elif workflow == 4:
            col_vals = data_rows.iloc[:, entity_ref_col_idx]
            should_export_entity_ref = any(pd.notna(v) and str(v).strip() not in ("", "nan", "None") for v in col_vals)
        else:
            should_export_entity_ref = False

    if entity_ref_col_idx is not None and not should_export_entity_ref:
        cols_to_keep = [i for i in range(completed_dc_df.shape[1]) if i != entity_ref_col_idx]
        data_rows = data_rows.iloc[:, cols_to_keep]
        completed_dc_df = completed_dc_df.iloc[:, cols_to_keep]

    final_csv_df = pd.concat([completed_dc_df.iloc[0:2], data_rows], ignore_index=True)
    
    output_file = os.path.join(output_dir, "DC.csv")
    final_csv_df.to_csv(output_file, index=False, header=False)
    print(f"\nSuccess! Exported fully validated metadata Excel to {excel_path}")
    print(f"Success! Exported fully validated metadata CSV to {output_file}")

    if opex_export:
        print("Generating OPEX metadata XML...")
        generate_opex_xml(completed_dc_df, output_dir, ref_folder_name)


class TextRedirector:
    def __init__(self, log_queue):
        self.log_queue = log_queue

    def write(self, str_val):
        self.log_queue.put(str_val)

    def flush(self):
        pass


class IngestGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Archival Metadata Creation Utility")
        
        # Automatically adjust window size depending on screen resolution (e.g. laptop vs desktop)
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        
        window_width = min(800, max(680, int(screen_width * 0.5)))
        window_height = min(950, max(550, int(screen_height * 0.85)))
        
        x = max(0, (screen_width - window_width) // 2)
        y = max(0, (screen_height - window_height) // 2)
        
        self.root.geometry(f"{window_width}x{window_height}+{x}+{y}")
        self.root.minsize(680, 550)
        
        # Load custom window icon/logo if available in project directory
        base_dir = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
        icon_relative_paths = [
            os.path.join("Images", "dominican-university-removebg-preview.png"),
            "logo.ico",
            "app_icon.ico",
            "icon.ico",
            "logo.png",
            "icon.png",
            "logo.gif",
            "dominican-university-removebg-preview.png"
        ]
        
        icon_candidates = []
        for path in icon_relative_paths:
            icon_candidates.append(os.path.join(base_dir, path))
            icon_candidates.append(path)
            
        self._icon_img = None
        for icon_file in icon_candidates:
            if os.path.exists(icon_file):
                try:
                    if icon_file.endswith(".ico"):
                        self.root.iconbitmap(icon_file)
                    else:
                        self._icon_img = tk.PhotoImage(file=icon_file)
                        self.root.iconphoto(True, self._icon_img)
                    break
                except Exception:
                    pass

        self.log_queue = queue.Queue()

        style = ttk.Style()
        if "clam" in style.theme_names():
            style.theme_use("clam")
            
        self.create_widgets()
        self.update_field_states()
        self.root.protocol("WM_DELETE_WINDOW", self.close_app)
        self.root.after(100, self.poll_log_queue)

    def create_widgets(self):
        # Canvas & Scrollbar container to ensure all elements remain fully viewable on any screen size (laptops & desktops)
        self.canvas = tk.Canvas(self.root, highlightthickness=0)
        self.scrollbar = ttk.Scrollbar(self.root, orient="vertical", command=self.canvas.yview)
        
        main_frame = ttk.Frame(self.canvas, padding="12")
        
        main_frame.bind(
            "<Configure>",
            lambda e: self.canvas.configure(scrollregion=self.canvas.bbox("all"))
        )
        
        self.canvas_window = self.canvas.create_window((0, 0), window=main_frame, anchor="nw")
        
        def _on_canvas_configure(event):
            self.canvas.itemconfig(self.canvas_window, width=event.width)
            
        self.canvas.bind("<Configure>", _on_canvas_configure)
        self.canvas.configure(yscrollcommand=self.scrollbar.set)
        
        def _on_mousewheel(event):
            self.canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            
        self.canvas.bind("<MouseWheel>", _on_mousewheel)
        main_frame.bind("<MouseWheel>", _on_mousewheel)
        
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Title / Banner Frame with Help Button on top right
        title_frame = ttk.Frame(main_frame)
        title_frame.pack(fill=tk.X, pady=(0, 10))

        if hasattr(self, '_icon_img') and self._icon_img:
            try:
                # Subsample 250x250 image for clean header logo display (~41x41)
                self._header_logo = self._icon_img.subsample(6, 6)
                logo_label = ttk.Label(title_frame, image=self._header_logo)
                logo_label.pack(side=tk.LEFT, padx=(0, 10))
            except Exception:
                pass

        title_label = ttk.Label(
            title_frame,
            text="Archival Metadata Utility Tool",
            font=("Segoe UI", 14, "bold")
        )
        title_label.pack(side=tk.LEFT, anchor=tk.W)

        btn_help = ttk.Button(
            title_frame,
            text="Help / Guide",
            command=self.open_help_guide
        )
        btn_help.pack(side=tk.RIGHT)

        # Option Selection Frame
        wf_frame = ttk.LabelFrame(main_frame, text=" 1. Select Option Method ", padding="10")
        wf_frame.pack(fill=tk.X, pady=(0, 10))

        self.workflow_var = tk.IntVar(value=1)

        rb1 = ttk.Radiobutton(
            wf_frame,
            text="Option 1: Digital Assets to be Ingested With Metadata",
            value=1,
            variable=self.workflow_var,
            command=self.update_field_states
        )
        rb1.pack(anchor=tk.W, pady=2)
        desc1 = ttk.Label(wf_frame, text="   Extracts technical metadata directly from source asset directory.", font=("Segoe UI", 9, "italic"), foreground="gray")
        desc1.pack(anchor=tk.W, pady=(0, 4))

        rb2 = ttk.Radiobutton(
            wf_frame,
            text="Option 2: Digital Assets Ingested Without Metadata (Adding IDs Later)",
            value=2,
            variable=self.workflow_var,
            command=self.update_field_states
        )
        rb2.pack(anchor=tk.W, pady=2)
        desc2 = ttk.Label(wf_frame, text="   Extracts technical metadata & merges Preservica Entity Reference IDs.", font=("Segoe UI", 9, "italic"), foreground="gray")
        desc2.pack(anchor=tk.W, pady=(0, 4))

        rb3 = ttk.Radiobutton(
            wf_frame,
            text="Option 3: Technical Metadata Existing",
            value=3,
            variable=self.workflow_var,
            command=self.update_field_states
        )
        rb3.pack(anchor=tk.W, pady=2)
        desc3 = ttk.Label(wf_frame, text="   Loads existing technical metadata file & merges Preservica Entity Reference IDs.", font=("Segoe UI", 9, "italic"), foreground="gray")
        desc3.pack(anchor=tk.W, pady=(0, 4))

        rb4 = ttk.Radiobutton(
            wf_frame,
            text="Option 4: Resume Prior Job / Validate Existing Metadata",
            value=4,
            variable=self.workflow_var,
            command=self.update_field_states
        )
        rb4.pack(anchor=tk.W, pady=2)
        desc4 = ttk.Label(wf_frame, text="   Resumes prior job & validates an existing metadata Excel file directly.", font=("Segoe UI", 9, "italic"), foreground="gray")
        desc4.pack(anchor=tk.W, pady=(0, 4))

        # OPEX Export Checkbox
        self.opex_var = tk.BooleanVar(value=False)
        cb_opex = ttk.Checkbutton(
            wf_frame,
            text="OPEX Export",
            variable=self.opex_var
        )
        cb_opex.pack(anchor=tk.W, pady=(4, 2))
        desc_opex = ttk.Label(
            wf_frame,
            text="   Generates Preservica OPEX metadata XML file(s) alongside standard export outputs.",
            font=("Segoe UI", 9, "italic"),
            foreground="gray"
        )
        desc_opex.pack(anchor=tk.W, pady=(0, 2))

        # Input Paths Frame
        paths_frame = ttk.LabelFrame(main_frame, text=" 2. Specify Input Paths ", padding="10")
        paths_frame.pack(fill=tk.X, pady=(0, 10))

        # Source Device Directory
        self.lbl_target = ttk.Label(paths_frame, text="Source Device Directory:")
        self.lbl_target.grid(row=0, column=0, sticky=tk.W, pady=4)
        self.entry_target = ttk.Entry(paths_frame, width=50)
        self.entry_target.grid(row=0, column=1, sticky=tk.EW, padx=5, pady=4)
        self.btn_target = ttk.Button(paths_frame, text="Browse Folder...", command=self.browse_target_dir)
        self.btn_target.grid(row=0, column=2, pady=4)

        # Preservica CSV
        self.lbl_preservica = ttk.Label(paths_frame, text="Preservica Export File:")
        self.lbl_preservica.grid(row=1, column=0, sticky=tk.W, pady=4)
        self.entry_preservica = ttk.Entry(paths_frame, width=50)
        self.entry_preservica.grid(row=1, column=1, sticky=tk.EW, padx=5, pady=4)
        self.btn_preservica = ttk.Button(paths_frame, text="Browse File...", command=self.browse_preservica_file)
        self.btn_preservica.grid(row=1, column=2, pady=4)

        # Existing Metadata
        self.lbl_existing = ttk.Label(paths_frame, text="Existing Metadata File:")
        self.lbl_existing.grid(row=2, column=0, sticky=tk.W, pady=4)
        self.entry_existing = ttk.Entry(paths_frame, width=50)
        self.entry_existing.grid(row=2, column=1, sticky=tk.EW, padx=5, pady=4)
        self.btn_existing = ttk.Button(paths_frame, text="Browse File...", command=self.browse_existing_metadata_file)
        self.btn_existing.grid(row=2, column=2, pady=4)

        # Working Metadata Workbook
        self.lbl_excel = ttk.Label(paths_frame, text="Metadata Template Excel:")
        self.lbl_excel.grid(row=3, column=0, sticky=tk.W, pady=4)
        self.entry_excel = ttk.Entry(paths_frame, width=50)
        self.entry_excel.insert(0, resolve_template_path("Metadata_Template.xlsx"))
        self.entry_excel.config(state=tk.DISABLED)
        self.entry_excel.grid(row=3, column=1, sticky=tk.EW, padx=5, pady=4)
        self.btn_excel = ttk.Button(paths_frame, text="Browse File...", command=self.browse_excel_file, state=tk.DISABLED)
        self.btn_excel.grid(row=3, column=2, pady=4)

        # Export Output Directory
        self.lbl_output_dir = ttk.Label(paths_frame, text="Export to:")
        self.lbl_output_dir.grid(row=4, column=0, sticky=tk.W, pady=4)
        self.entry_output_dir = ttk.Entry(paths_frame, width=50)
        self.entry_output_dir.grid(row=4, column=1, sticky=tk.EW, padx=5, pady=4)
        self.btn_output_dir = ttk.Button(paths_frame, text="Browse Folder...", command=self.browse_output_dir)
        self.btn_output_dir.grid(row=4, column=2, pady=4)

        paths_frame.columnconfigure(1, weight=1)

        # Action Frame
        action_frame = ttk.Frame(main_frame)
        action_frame.pack(fill=tk.X, pady=(0, 10))

        self.btn_run = ttk.Button(
            action_frame,
            text="Start Process",
            command=self.start_process_thread
        )
        self.btn_run.pack(side=tk.LEFT, padx=5)

        self.btn_exit = ttk.Button(
            action_frame,
            text="Exit",
            command=self.close_app
        )
        self.btn_exit.pack(side=tk.LEFT, padx=5)

        self.lbl_status = ttk.Label(action_frame, text="Ready", font=("Segoe UI", 10, "bold"), foreground="green")
        self.lbl_status.pack(side=tk.LEFT, padx=10)

        # Output / Log Frame
        log_frame = ttk.LabelFrame(main_frame, text=" 3. Process Output Log ", padding="10")
        log_frame.pack(fill=tk.BOTH, expand=True)

        self.log_text = scrolledtext.ScrolledText(log_frame, height=10, font=("Consolas", 9), wrap=tk.WORD)
        self.log_text.pack(fill=tk.BOTH, expand=True, pady=(0, 5))

        log_btn_frame = ttk.Frame(log_frame)
        log_btn_frame.pack(fill=tk.X)

        self.btn_print_log = ttk.Button(
            log_btn_frame,
            text="Print Output Log (.txt)",
            command=self.save_output_log
        )
        self.btn_print_log.pack(side=tk.RIGHT)

        def _bind_mousewheel(widget):
            if not isinstance(widget, scrolledtext.ScrolledText):
                widget.bind("<MouseWheel>", _on_mousewheel, add="+")
                for child in widget.winfo_children():
                    _bind_mousewheel(child)
        _bind_mousewheel(main_frame)

    def open_help_guide(self):
        """
        Opens the Beginner's Guide markdown document in the default editor / viewer.
        """
        guide_file = "Beginners-Guide_Automated-Metadata-Extraction-Tool.md"
        abs_path = os.path.abspath(guide_file)
        if not os.path.exists(abs_path):
            script_dir = os.path.dirname(os.path.abspath(__file__))
            candidate = os.path.join(script_dir, guide_file)
            if os.path.exists(candidate):
                abs_path = candidate

        if os.path.exists(abs_path):
            try:
                if hasattr(os, 'startfile'):
                    os.startfile(abs_path)
                else:
                    import subprocess
                    subprocess.Popen(['open' if sys.platform == 'darwin' else 'xdg-open', abs_path])
                print(f"Opened Help Guide: {abs_path}")
            except Exception as e:
                messagebox.showerror("Error Opening Guide", f"Could not open guide file:\n{abs_path}\n\nError: {e}")
        else:
            messagebox.showinfo("Guide File Not Found", f"Help guide file not found:\n{guide_file}")

    def find_created_metadata_file(self):
        """
        Locates the metadata Excel workbook or CSV file created in the export directory.
        """
        output_dir = self.entry_output_dir.get().strip()
        existing_meta = self.entry_existing.get().strip()
        
        target_file = None
        
        # 1. Search output_dir for created metadata files
        if output_dir and os.path.exists(output_dir):
            excel_files = [os.path.join(output_dir, f) for f in os.listdir(output_dir) if f.endswith("_Metadata.xlsx") and not f.startswith("~$")]
            if not excel_files:
                excel_files = [os.path.join(output_dir, f) for f in os.listdir(output_dir) if f.endswith(".xlsx") and not f.startswith("~$")]
            
            if excel_files:
                excel_files.sort(key=os.path.getmtime, reverse=True)
                target_file = excel_files[0]
            elif os.path.exists(os.path.join(output_dir, "DC.csv")):
                target_file = os.path.join(output_dir, "DC.csv")

        # 2. Fallback to existing_meta if specified and exists
        if not target_file and existing_meta and os.path.exists(existing_meta):
            target_file = existing_meta

        return target_file

    def open_metadata_file(self):
        """
        Opens the metadata Excel workbook or CSV file created in the export directory.
        """
        target_file = self.find_created_metadata_file()
        if target_file and os.path.exists(target_file):
            try:
                if hasattr(os, 'startfile'):
                    os.startfile(target_file)
                else:
                    import subprocess
                    subprocess.Popen(['open' if sys.platform == 'darwin' else 'xdg-open', target_file])
                print(f"Opened metadata file: {target_file}")
            except Exception as e:
                messagebox.showerror("Error Opening File", f"Could not open metadata file:\n{target_file}\n\nError: {e}")
        else:
            messagebox.showinfo("File Not Found", "No created metadata file was found in the export directory.\n\nPlease run the process first or select an export directory containing the metadata file.")

    def save_output_log(self):
        log_content = self.log_text.get("1.0", tk.END).strip()
        if not log_content:
            messagebox.showinfo("Log Empty", "There is no log output to save.")
            return

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_filename = f"Process_Log_{timestamp}.txt"

        init_dir = self.entry_output_dir.get().strip() if hasattr(self, 'entry_output_dir') else ""
        if not init_dir or not os.path.exists(init_dir):
            init_dir = os.getcwd()

        file_path = filedialog.asksaveasfilename(
            title="Save Process Output Log",
            initialdir=init_dir,
            initialfile=default_filename,
            defaultextension=".txt",
            filetypes=[("Text Files", "*.txt"), ("Log Files", "*.log"), ("All Files", "*.*")]
        )
        if file_path:
            try:
                with open(file_path, "w", encoding="utf-8") as f:
                    f.write(log_content + "\n")
                messagebox.showinfo("Log Saved", f"Process output log saved successfully to:\n{file_path}")
            except Exception as e:
                messagebox.showerror("Save Error", f"Could not save log file:\n{e}")

    def update_field_states(self):
        wf = self.workflow_var.get()
        
        # Target Directory: required for 1 & 2
        if wf in (1, 2):
            self.entry_target.config(state=tk.NORMAL)
            self.btn_target.config(state=tk.NORMAL)
            self.lbl_target.config(foreground="black")
        else:
            self.entry_target.config(state=tk.DISABLED)
            self.btn_target.config(state=tk.DISABLED)
            self.lbl_target.config(foreground="gray")

        # Preservica CSV: required for 2 & 3
        if wf in (2, 3):
            self.entry_preservica.config(state=tk.NORMAL)
            self.btn_preservica.config(state=tk.NORMAL)
            self.lbl_preservica.config(foreground="black")
        else:
            self.entry_preservica.config(state=tk.DISABLED)
            self.btn_preservica.config(state=tk.DISABLED)
            self.lbl_preservica.config(foreground="gray")

        # Existing Metadata: required for 3, optional for 4
        if wf in (3, 4):
            self.entry_existing.config(state=tk.NORMAL)
            self.btn_existing.config(state=tk.NORMAL)
            self.lbl_existing.config(foreground="black")
            if wf == 4:
                existing_path = self.entry_existing.get().strip()
                if existing_path and os.path.exists(existing_path):
                    parent_dir = os.path.dirname(os.path.abspath(existing_path))
                    if parent_dir and os.path.abspath(parent_dir) != os.path.abspath("."):
                        self.entry_output_dir.delete(0, tk.END)
                        self.entry_output_dir.insert(0, os.path.normpath(parent_dir))
        else:
            self.entry_existing.config(state=tk.DISABLED)
            self.btn_existing.config(state=tk.DISABLED)
            self.lbl_existing.config(foreground="gray")

    def browse_target_dir(self):
        path = filedialog.askdirectory(title="Select Source Device Directory")
        if path:
            self.entry_target.delete(0, tk.END)
            self.entry_target.insert(0, os.path.normpath(path))

    def browse_preservica_file(self):
        path = filedialog.askopenfilename(
            title="Select Preservica Export File",
            filetypes=[("CSV and Excel files", "*.csv *.xlsx *.xls"), ("CSV files", "*.csv"), ("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if path:
            self.entry_preservica.delete(0, tk.END)
            self.entry_preservica.insert(0, os.path.normpath(path))

    def browse_existing_metadata_file(self):
        path = filedialog.askopenfilename(
            title="Select Existing Technical Metadata File",
            filetypes=[("Excel and CSV files", "*.xlsx *.xls *.csv"), ("Excel files", "*.xlsx *.xls"), ("CSV files", "*.csv"), ("All files", "*.*")]
        )
        if path:
            norm_p = os.path.normpath(path)
            self.entry_existing.delete(0, tk.END)
            self.entry_existing.insert(0, norm_p)
            
            # Automatically set 'Export to:' to the folder containing the selected existing metadata file
            parent_dir = os.path.dirname(os.path.abspath(norm_p))
            if parent_dir and os.path.abspath(parent_dir) != os.path.abspath("."):
                self.entry_output_dir.delete(0, tk.END)
                self.entry_output_dir.insert(0, os.path.normpath(parent_dir))

    def browse_excel_file(self):
        path = filedialog.askopenfilename(
            title="Select Metadata Template Excel Workbook",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if path:
            self.entry_excel.delete(0, tk.END)
            self.entry_excel.insert(0, os.path.normpath(path))

    def browse_output_dir(self):
        path = filedialog.askdirectory(title="Select Export Output Directory")
        if path:
            self.entry_output_dir.delete(0, tk.END)
            self.entry_output_dir.insert(0, os.path.normpath(path))

    def close_app(self):
        try:
            self.root.destroy()
        except Exception:
            pass
        sys.exit(0)

    def poll_log_queue(self):
        while not self.log_queue.empty():
            msg = self.log_queue.get_nowait()
            self.log_text.insert(tk.END, msg)
            self.log_text.see(tk.END)
        self.root.after(100, self.poll_log_queue)

    def start_process_thread(self):
        wf = self.workflow_var.get()
        target_dir = self.entry_target.get().strip()
        preservica_csv = self.entry_preservica.get().strip()
        existing_meta = self.entry_existing.get().strip()
        template_path = self.entry_excel.get().strip() or "Metadata_Template.xlsx"
        output_dir = self.entry_output_dir.get().strip()

        if wf == 4:
            if existing_meta:
                if not os.path.exists(existing_meta):
                    messagebox.showerror("Path Error", f"Existing Metadata File does not exist:\n{existing_meta}")
                    return
                if not output_dir:
                    output_dir = os.path.dirname(os.path.abspath(existing_meta))
                    self.entry_output_dir.delete(0, tk.END)
                    self.entry_output_dir.insert(0, os.path.normpath(output_dir))

        # Validation
        if not output_dir:
            messagebox.showerror("Input Error", "Please select an Export Output Directory or Existing Metadata File.")
            return

        if os.path.abspath(output_dir) == os.path.abspath("."):
            messagebox.showerror("Output Directory Error", "Export directory cannot be the root directory. Please select a different export output directory.")
            return

        if wf in (1, 2):
            if not target_dir:
                messagebox.showerror("Input Error", f"Please select a Source Device Directory for Option {wf}")
                return
            if not os.path.exists(target_dir):
                messagebox.showerror("Path Error", f"Source Device Directory does not exist:\n{target_dir}")
                return

        if wf in (2, 3):
            if not preservica_csv:
                messagebox.showerror("Input Error", f"Please select a Preservica Export File for Option {wf}")
                return
            if not os.path.exists(preservica_csv):
                messagebox.showerror("Path Error", f"Preservica Export File does not exist:\n{preservica_csv}")
                return

        if wf == 3:
            if not existing_meta:
                messagebox.showerror("Input Error", "Please select an Existing Metadata File for Option 3")
                return
            if not os.path.exists(existing_meta):
                messagebox.showerror("Path Error", f"Existing Metadata File does not exist:\n{existing_meta}")
                return

        if wf == 4:
            if existing_meta and not os.path.exists(existing_meta):
                messagebox.showerror("Path Error", f"Existing Metadata File does not exist:\n{existing_meta}")
                return

        template_path = resolve_template_path(template_path)
        if not os.path.exists(template_path):
            messagebox.showerror("Path Error", f"Metadata Template Workbook does not exist:\n{template_path}")
            return

        self.btn_run.config(state=tk.DISABLED)
        self.lbl_status.config(text="Processing...", foreground="orange")
        self.log_text.delete("1.0", tk.END)

        def worker():
            old_stdout = sys.stdout
            sys.stdout = TextRedirector(self.log_queue)
            try:
                def gui_resume_handler(excel_path):
                    event = threading.Event()
                    user_choice = [False]
                    def show_resume_popup():
                        filename = os.path.basename(excel_path)
                        res = messagebox.askyesno(
                            "Existing Metadata Workbook Found",
                            f"An existing metadata workbook was found in the export directory:\n{filename}\n\n"
                            f"Would you like to RESUME using this existing workbook?\n\n"
                            f"• Click YES to preserve your existing DC edits and proceed to validation.\n"
                            f"• Click NO to overwrite with a fresh template."
                        )
                        user_choice[0] = res
                        event.set()
                    self.root.after(0, show_resume_popup)
                    event.wait()
                    return user_choice[0]

                def gui_pause_handler(msg):
                    event = threading.Event()
                    def show_popup():
                        dialog = tk.Toplevel(self.root)
                        dialog.title("ACTION REQUIRED: Edit Metadata")
                        dialog.geometry("560x220")
                        dialog.transient(self.root)
                        dialog.grab_set()

                        if hasattr(self, '_icon_img') and self._icon_img:
                            try:
                                dialog.iconphoto(False, self._icon_img)
                            except Exception:
                                pass

                        # Center relative to root window
                        x = self.root.winfo_x() + max(0, (self.root.winfo_width() - 560) // 2)
                        y = self.root.winfo_y() + max(0, (self.root.winfo_height() - 220) // 2)
                        dialog.geometry(f"+{x}+{y}")

                        dlg_frame = ttk.Frame(dialog, padding="15")
                        dlg_frame.pack(fill=tk.BOTH, expand=True)

                        lbl_text = f"{msg}\n\nClick OK once you have saved and closed the Excel file to proceed with validation & export."
                        lbl = ttk.Label(dlg_frame, text=lbl_text, wraplength=520, font=("Segoe UI", 9))
                        lbl.pack(anchor=tk.W, fill=tk.BOTH, expand=True, pady=(0, 15))

                        btn_frame = ttk.Frame(dlg_frame)
                        btn_frame.pack(fill=tk.X, side=tk.BOTTOM)

                        def open_file_action():
                            target_p = None
                            if "Please open " in msg:
                                try:
                                    target_p = msg.split("Please open ")[1].split(", complete")[0].strip()
                                except Exception:
                                    target_p = None
                            if not target_p or not os.path.exists(target_p):
                                target_p = self.find_created_metadata_file()

                            if target_p and os.path.exists(target_p):
                                try:
                                    if hasattr(os, 'startfile'):
                                        os.startfile(target_p)
                                    else:
                                        import subprocess
                                        subprocess.Popen(['open' if sys.platform == 'darwin' else 'xdg-open', target_p])
                                    print(f"Opened metadata file: {target_p}")
                                except Exception as e:
                                    messagebox.showerror("Error Opening File", f"Could not open metadata file:\n{target_p}\n\nError: {e}")
                            else:
                                messagebox.showinfo("File Not Found", "No metadata file was found to open.")

                        def ok_action():
                            event.set()
                            dialog.destroy()

                        dialog.protocol("WM_DELETE_WINDOW", ok_action)

                        btn_open = ttk.Button(btn_frame, text="Open File", command=open_file_action)
                        btn_open.pack(side=tk.LEFT, padx=(0, 10))

                        btn_ok = ttk.Button(btn_frame, text="OK", command=ok_action)
                        btn_ok.pack(side=tk.LEFT)

                    self.root.after(0, show_popup)
                    event.wait()

                opex_export = self.opex_var.get()

                if wf == 1:
                    process_metadata(source_folder=target_dir, template_path=template_path, output_dir=output_dir, on_pause_callback=gui_pause_handler, on_resume_callback=gui_resume_handler, workflow=1, opex_export=opex_export)
                elif wf == 2:
                    process_metadata(source_folder=target_dir, template_path=template_path, preservica_csv=preservica_csv, output_dir=output_dir, on_pause_callback=gui_pause_handler, on_resume_callback=gui_resume_handler, workflow=2, opex_export=opex_export)
                elif wf == 3:
                    process_metadata(template_path=template_path, preservica_csv=preservica_csv, existing_metadata=existing_meta, output_dir=output_dir, on_pause_callback=gui_pause_handler, on_resume_callback=gui_resume_handler, workflow=3, opex_export=opex_export)
                elif wf == 4:
                    process_metadata(existing_metadata=existing_meta, template_path=template_path, output_dir=output_dir, on_pause_callback=gui_pause_handler, resume=True, workflow=4, opex_export=opex_export)
                
                self.root.after(0, lambda: self.lbl_status.config(text="Completed Successfully", foreground="green"))
                self.root.after(0, lambda: messagebox.showinfo("Process Complete", f"Metadata process completed successfully!\nFiles exported to: {output_dir}"))
            except Exception as e:
                print(f"\nERROR during process execution: {e}")
                self.root.after(0, lambda: self.lbl_status.config(text="Failed with Error", foreground="red"))
                self.root.after(0, lambda: messagebox.showerror("Process Error", f"An error occurred:\n{e}"))
            finally:
                sys.stdout = old_stdout
                self.root.after(0, lambda: self.btn_run.config(state=tk.NORMAL))

        threading.Thread(target=worker, daemon=True).start()


def launch_gui():
    root = tk.Tk()
    app = IngestGUI(root)
    root.mainloop()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Automated Ingest Utility for Preservica options.")
    parser.add_argument("--option", "--workflow", type=int, choices=[1, 2, 3, 4], required=False, dest="workflow",
                        help="Select Option 1 (Metadata Included), 2 (Adding IDs later), 3 (Existing Technical Metadata), or 4 (Resume Prior Job / Validate Existing Metadata)")
    parser.add_argument("--target-dir", type=str, 
                        help="Path to the directory containing the digital assets (Required for options 1 and 2)")
    parser.add_argument("--template-path", "--excel-path", type=str, default="Metadata_Template.xlsx", 
                        help="Path to the Metadata_Template.xlsx file")
    parser.add_argument("--preservica-csv", type=str, 
                        help="Path to Preservica export CSV (Required if running Option 2 or 3)")
    parser.add_argument("--existing-metadata", type=str, 
                        help="Path to the existing technical metadata Excel/CSV file (Required for option 3)")
    parser.add_argument("--output-dir", type=str, 
                        help="Directory path where the metadata Excel copy and DC.csv will be saved (Required, must be different from root directory)")
    parser.add_argument("--resume", action="store_true", help="Resume processing using existing metadata workbook if found in export directory")
    parser.add_argument("--opex", "--opex-export", action="store_true", dest="opex_export", help="Generate OPEX metadata XML file(s) during export")
    parser.add_argument("--gui", action="store_true", help="Force launch Graphical User Interface (GUI)")
    
    args = parser.parse_args()
    
    if args.gui or args.workflow is None:
        launch_gui()
    else:
        print(f"Initializing Option {args.workflow}...")
        
        if not args.output_dir or os.path.abspath(args.output_dir) == os.path.abspath("."):
            print("Error: Export directory (--output-dir) is required and cannot be the root directory.")
        elif args.workflow == 1:
            if not args.target_dir:
                print("Error: Option 1 requires --target-dir.")
            else:
                process_metadata(source_folder=args.target_dir, template_path=args.template_path, output_dir=args.output_dir, resume=args.resume, workflow=1, opex_export=args.opex_export)
        
        elif args.workflow == 2:
            if not args.target_dir:
                print("Error: Option 2 requires --target-dir.")
            elif not args.preservica_csv:
                print("Error: Option 2 requires the --preservica-csv argument to merge Entity IDs.")
            else:
                process_metadata(source_folder=args.target_dir, template_path=args.template_path, preservica_csv=args.preservica_csv, output_dir=args.output_dir, resume=args.resume, workflow=2, opex_export=args.opex_export)
        
        elif args.workflow == 3:
            if not args.existing_metadata:
                print("Error: Option 3 requires --existing-metadata to load existing technical metadata.")
            elif not args.preservica_csv:
                print("Error: Option 3 requires the --preservica-csv argument to merge Entity IDs.")
            else:
                process_metadata(template_path=args.template_path, preservica_csv=args.preservica_csv, existing_metadata=args.existing_metadata, output_dir=args.output_dir, resume=args.resume, workflow=3, opex_export=args.opex_export)
        
        elif args.workflow == 4:
            out_dir = args.output_dir
            if not out_dir and args.existing_metadata and os.path.exists(args.existing_metadata):
                out_dir = os.path.dirname(os.path.abspath(args.existing_metadata))
            if not out_dir or os.path.abspath(out_dir) == os.path.abspath("."):
                print("Error: Export directory (--output-dir) is required or derived from --existing-metadata, and cannot be the root directory.")
            else:
                process_metadata(existing_metadata=args.existing_metadata, template_path=args.template_path, output_dir=out_dir, resume=True, workflow=4, opex_export=args.opex_export)